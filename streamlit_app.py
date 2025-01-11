import streamlit as st
import pandas as pd
import openpyxl
from io import BytesIO
from openpyxl.utils import get_column_letter
import itertools
import re
from notion_client import Client
import webbrowser

st.set_page_config(page_icon="📊", page_title="Price list Analysis")

st.title("Price list Analysis")

def extract_sku_details(uploaded_file):
    # Load the Excel file
    xls = pd.ExcelFile(uploaded_file)
    product_df = pd.read_excel(xls, sheet_name='Product')
    
    # Replace any '0', 'WeightAdjustment', or empty values in the 'SKU' column withr NaN for forward filling
    product_df['SKU'] = product_df['SKU'].replace([0, '0', '', 'WeightAdjustment'], pd.NA)
    
    # Get the initial SKU value from cell Q2 (assuming it's in the second row)
    initial_sku = product_df.loc[1, 'SKU']
    
    # Forward-fill the SKU column
    product_df['SKU'] = product_df['SKU'].ffill()
    
    # Fill any remaining NaN values (at the beginning) with the initial SKU
    product_df['SKU'] = product_df['SKU'].fillna(initial_sku)
    
    # Variables to store data
    result_data = []
    sku = None
    price = None
    product_cost = None

    # Iterate through the DataFrame row by row
    for idx, row in product_df.iterrows():
        # Get the current SKU, which has been forward-filled
        current_sku = row['SKU']

        # If a new SKU is detected, update the price and product_cost from the same row
        if current_sku != sku:
            # Save the previous SKU's data if any
            if sku is not None:
                result_data.append({
                    'SKU': sku,
                    'Price': price,
                    'ProductCost': product_cost,
                })

            price = row['Price']  # Replace 'Price' with the actual column name for price
            product_cost = row['ProductCost']  # Replace 'ProductCost' with the actual column name

        # Update the current SKU
        sku = current_sku

    # Store data for the last SKU if there's any remaining data
    if sku is not None:
        result_data.append({
            'SKU': sku,
            'Price': price,
            'ProductCost': product_cost,
        })

    # Convert the results to a DataFrame
    result_df = pd.DataFrame(result_data)

    return result_df

def generate_apn_list(result_df):
    # List to store APN records
    apn_records = []

    # Iterate over each row in result_df
    for idx, row in result_df.iterrows():
        sku = row['SKU']
        price = row['Price']
        product_cost = row['ProductCost']

        # Split the sizes and colors into lists
        sizes = split_sizes(str(row['Sizes'])) if pd.notna(row['Sizes']) else []
        colors = str(row['Colors']).split(',') if pd.notna(row['Colors']) else []

        # Generate all combinations of sizes and colors
        if sizes and colors:
            combinations = itertools.product(sizes, colors)
        elif sizes:
            combinations = [(size, None) for size in sizes]
        elif colors:
            combinations = [(None, color) for color in colors]
        else:
            combinations = [(None, None)]

        # Create a record for each combination
        for size, color in combinations:
            apn_records.append({
                'SKU': sku,
                'Size': size,
                'Color': color,
                'Price': price,
                'ProductCost': product_cost
            })

    # Convert the APN records to a DataFrame
    apn_df = pd.DataFrame(apn_records)

    return apn_df

def split_sizes(size_string):
    # Pattern to match commas that are outside of parentheses
    pattern = r',\s*(?![^()]*\))'
    sizes = re.split(pattern, size_string)
    return [s.strip() for s in sizes]

# Define a function to apply conditional formatting
def color_difference(val):
    if isinstance(val, (int, float)):  # Ensure it's a number
        if val > 0:
            return 'color: green'
        elif val < 0:
            return 'color: red'
    return ''  # Default style

# Notion API Token
NOTION_TOKEN = st.secrets["notion"]["NOTION_TOKEN"]
DATABASE_ID = st.secrets["notion"]["DATABASE_ID"]

# Initialize Notion Client
notion = Client(auth=NOTION_TOKEN)

def fetch_notion_data(database_id):
    # Fetch the Notion database structure
    database = notion.databases.retrieve(database_id=database_id)
    properties = database.get("properties", {})

    # Extract property names
    ordered_columns = list(properties.keys())

    # Ensure 'Name' column is first if it exists
    if "Name" in ordered_columns:
        ordered_columns.remove("Name")
        ordered_columns = ["Name"] + ordered_columns

    # Fetch the database content
    results = notion.databases.query(database_id=database_id).get("results")

    data = []
    for page in results:
        row = {}
        page_id = page["id"].replace("-", "")  # Remove hyphens for URL formatting
        notion_url = f"https://www.notion.so/{page_id}"  # Notion page URL

        for key in ordered_columns:
            prop = page["properties"].get(key, {})
            prop_type = prop.get("type")
            prop_value = prop.get(prop_type)

            # Check for title type
            if prop_type == "title" and prop_value and len(prop_value) > 0:
                title = prop_value[0].get("text", {}).get("content", "")
                # Create a markdown link to the Notion page
                row[key] = f"[{title}]({notion_url})"
            # Check for select/multi-select types
            elif prop_type in ["select", "multi_select"] and prop_value:
                if isinstance(prop_value, list):
                    row[key] = ", ".join([item.get("name", "") for item in prop_value])
                else:
                    row[key] = prop_value.get("name", "")
            else:
                row[key] = ""
        data.append(row)

    # Convert to DataFrame and reorder columns
    df = pd.DataFrame(data)
    df = df[ordered_columns]  # Ensure 'Name' is first

    return df

def fetch_brand_names(database_id):
    # Query the Notion database
    results = notion.databases.query(database_id=database_id).get("results")

    brand_names = set()
    for page in results:
        properties = page["properties"]
        # Replace 'Brand' with the actual property name for brands in your database
        brand_property = properties.get("Brand", {})
        brand_type = brand_property.get("type")
        brand_value = brand_property.get(brand_type)

        # Handle select or multi_select fields
        if brand_type == "select" and brand_value:
            brand_names.add(brand_value.get("name", ""))
        elif brand_type == "multi_select" and brand_value:
            for item in brand_value:
                brand_names.add(item.get("name", ""))

    return sorted(list(brand_names))

############################## PRICELIST STREAMLIT APP ####################################################

# Fetch data from Notion
df = fetch_notion_data(DATABASE_ID)

# Extract only the text inside square brackets from the 'Name' column
if 'Name' in df.columns:
    # Use regex to extract the text inside square brackets
    brand_list = df['Name'].dropna().apply(lambda x: re.search(r'\[(.*?)\]', x).group(1) if re.search(r'\[(.*?)\]', x) else x).unique().tolist()
    brand_list.sort()  # Optional: Sort the brand names
else:
    brand_list = ["No Brands Found"]

# Populate the selectbox with brand names
selected_brand = st.selectbox("Select Brand", brand_list)

with st.expander("Paramaters", expanded=True):
    # Check if 'RRP' column exists in the data
    if 'RRP' in df.columns:
        # Filter for the selected brand
        selected_row = df[df['Name'].str.contains(selected_brand, na=False, case=False)]
        
        # Check if RRP is 'Yes' for the selected brand
        is_rrp_checked = selected_row['RRP'].str.lower().eq('yes').any()
        
        # Display the checkbox
        RRP_required = st.checkbox("RRP", value=is_rrp_checked, disabled=True)
    else:
        st.warning("RRP column not found in the database.")

    # Get the Notion URL for the selected brand
    selected_row = df[df['Name'].str.contains(selected_brand, na=False, case=False)]

    if not selected_row.empty:
        # Extract the original Notion link from the markdown format
        notion_link = selected_row['Name'].iloc[0]
        notion_url_match = re.search(r'\((.*?)\)', notion_link)
        notion_url = notion_url_match.group(1) if notion_url_match else None

        # Create a Streamlit button to open the Notion page
        if notion_url:
            if st.button("🔗 Open Notion Page"):
                webbrowser.open_new_tab(notion_url)
        else:
            st.warning("No Notion URL found for the selected brand.")
    else:
        st.warning("No data found for the selected brand.")

# if st.button("Load Notion Data"):
#     df = fetch_notion_data(DATABASE_ID)
#     st.dataframe(df)

tab1, tab2, tab3 = st.tabs(["Preliminary Price list", "Old Price list", "Compare Old & New Price list"])

with tab1:
    # Load Excel file
    file = st.file_uploader("Upload file", type=["xlsx"])

    # Dropdown options for column selection
    letters = [""] + [chr(i) for i in range(65, 91)]  # Adds an empty option to ['A', 'B', 'C', ..., 'Z']

    # Dropdown for SKU column selection
    sku_column = st.selectbox("Select the SKU Column (A-Z):", letters)
    price_column = st.selectbox("Select the Product Cost Column (A-Z):", letters)

    if RRP_required is True:
        rrp_column = st.selectbox("Select the RRP Column (A-Z):", letters)

    if file is not None:
        try:
            # Load the Excel file
            df = pd.read_excel(file)

            if sku_column and price_column:
                # Map letter to column index
                sku_col_index = ord(sku_column) - 65  # Convert letter to zero-based index
                price_col_index = ord(price_column) - 65  # Convert letter to zero-based index
                
                if RRP_required is True:
                    RRP_col_index = ord(rrp_column) - 65  # Convert letter to zero-based index

                if 0 <= sku_col_index < len(df.columns) and 0 <= price_col_index < len(df.columns):
                    sku_col_name = df.columns[sku_col_index]
                    price_col_name = df.columns[price_col_index]

                    if is_rrp_checked:
                        RRP_col_name = df.columns[RRP_col_index]

                    # Ensure the price column is numeric
                    df[price_col_name] = pd.to_numeric(df[price_col_name], errors="coerce")

                    # Extract unique SKUs and their counts
                    sku_counts = df[sku_col_name].value_counts()

                    # Add new tab to the Excel file
                    file.seek(0)  # Reset file pointer
                    workbook = openpyxl.load_workbook(file)
                    new_sheet_name = "New Pricelist"
                    
                    if new_sheet_name not in workbook.sheetnames:
                        sheet = workbook.create_sheet(new_sheet_name)
                        
                        # Write headers
                        sheet.cell(row=1, column=1, value="SKU")  # Header for Column A
                        sheet.cell(row=1, column=2, value="Count")  # Header for Column B
                        sheet.cell(row=1, column=3, value="Average Cost")  # Header for Column C
                        sheet.cell(row=1, column=4, value="Price Range")  # Header for Column D
                        
                        # Write unique SKUs, their counts, average prices, and ranges
                        for i, (sku, count) in enumerate(sku_counts.items(), start=2):
                            sheet.cell(row=i, column=1, value=sku)  # SKU
                            sheet.cell(row=i, column=2, value=count)  # Count

                            # Filter rows for the current SKU and drop NaN prices
                            sku_data = df.loc[df[sku_col_name] == sku, price_col_name].dropna()

                            # Calculate the average price for this SKU
                            avg_price = sku_data.mean()
                            sheet.cell(row=i, column=3, value=avg_price)  # Average price

                            # Calculate the price range for this SKU
                            if not sku_data.empty:
                                max_price = sku_data.max()
                                min_price = sku_data.min()
                                price_range = max_price - min_price
                            else:
                                price_range = None
                            sheet.cell(row=i, column=4, value=price_range)  # Price range
                        
                        st.success(f"New sheet '{new_sheet_name}' added to the file.")
                    else:
                        st.warning(f"Sheet '{new_sheet_name}' already exists.")
                    
                    # Save the modified Excel file for download
                    output = BytesIO()
                    workbook.save(output)
                    output.seek(0)
                    
                    st.download_button(
                        label= f"Download {selected_brand} NEW Pricelist file",
                        data=output,
                        file_name= f"{selected_brand} NEW Pricelist file.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    )
                else:
                    st.error(f"Invalid column letter '{sku_column}' or '{price_column}'. No corresponding column found in the file.")
            else:
                st.warning("Please select valid SKU and Price columns.")
        except Exception as e:
            st.error(f"Error reading or processing the file: {e}")

with tab2:
    st.write("Upload the product data file to extract sizes and colors for each SKU.")
    # File upload for SKU extraction
    uploaded_file = st.file_uploader("Choose an Excel file", type=['xlsx'])

    if uploaded_file is not None:
        st.write("File uploaded successfully. Processing...")
        
        # Extract SKU details
        result = extract_sku_details(uploaded_file)
        st.write("Extracted SKU Details (showing first 100 rows):")
        st.write(result.head(100))
        
        # Provide download button for the SKU details
        output = BytesIO()
        result.to_excel(output, index=False, engine='openpyxl')
        output.seek(0)

        st.download_button(
            label= f"Download {selected_brand} OLD Pricelist file",
            data=output,
            file_name= f"{selected_brand} OLD Pricelist file.xlsx",
            key="old_pricelist",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

with tab3:
    
    # File uploaders
    preliminary_file = st.file_uploader("Upload the Preliminary Price List", type=['xlsx'])
    old_file = st.file_uploader("Upload the Old Price List", type=['xlsx'])

    if preliminary_file and old_file:
        try:
            # Load the files into DataFrames
            preliminary_df = pd.read_excel(preliminary_file)
            old_df = pd.read_excel(old_file)

            # Normalize the SKU column
            preliminary_df['SKU'] = preliminary_df['SKU'].astype(str).str.strip().str.upper()
            old_df['SKU'] = old_df['SKU'].astype(str).str.strip().str.upper()

            # Sort the dataframes
            preliminary_df = preliminary_df.sort_values(by='SKU')
            old_df = old_df.sort_values(by='SKU')

            # Assuming both files have 'SKU' and 'Price' columns
            # Rename price columns for clarity
            preliminary_df = preliminary_df.rename(columns={"ProductCost": "New Price"})
            old_df = old_df.rename(columns={"ProductCost": "Old Price"})

            # Merge the DataFrames on SKU with an outer join to include all SKUs
            merged_df = pd.merge(preliminary_df, old_df, on='SKU', how='outer')

            # Add a 'Difference' column
            merged_df['Difference'] = merged_df['New Price'] - merged_df['Old Price']

            # Add a 'Status' column based on the conditions
            def determine_status(row):
                if row['New Price'] == 'NA' and pd.notna(row['Old Price']):
                    return "Discontinued"
                elif row['Old Price'] == 'NA' and pd.notna(row['New Price']):
                    return "New Product"
                else:
                    return "Unchanged"

            # Fill NaN values for clarity
            merged_df = merged_df.fillna({
                'New Price': 'NA',
                'Old Price': 'NA',
                'Difference': 'NA'
            })

            # Apply the status logic
            merged_df['Status'] = merged_df.apply(determine_status, axis=1)

            # Apply conditional formatting to the 'Difference' column
            styled_df = merged_df.style.applymap(color_difference, subset=['Difference'])

            # Display the styled DataFrame
            st.write("Comparison of Prices (showing first 100 rows):")
            st.dataframe(styled_df)

            # Provide a download button for the new Excel file
            output = BytesIO()
            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                merged_df.to_excel(writer, index=False, sheet_name="Price Comparison")
                workbook = writer.book
                sheet = writer.sheets["Price Comparison"]

                # Get the column number for "Difference" (1-based index)
                difference_col = merged_df.columns.get_loc("Difference") + 1
                max_row = len(merged_df) + 1  # Include header row

                # Define green and red fills for conditional formatting
                green_fill = openpyxl.styles.PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                red_fill = openpyxl.styles.PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

                # Apply conditional formatting for positive (green) and negative (red) differences
                sheet.conditional_formatting.add(
                    f"{get_column_letter(difference_col)}2:{get_column_letter(difference_col)}{max_row}",
                    openpyxl.formatting.rule.CellIsRule(operator="greaterThan", formula=["0"], fill=green_fill)
                )
                sheet.conditional_formatting.add(
                    f"{get_column_letter(difference_col)}2:{get_column_letter(difference_col)}{max_row}",
                    openpyxl.formatting.rule.CellIsRule(operator="lessThan", formula=["0"], fill=red_fill)
                )

            output.seek(0)
            st.download_button(
                label="Download Price Comparison File with Formatting",
                data=output,
                file_name="price_comparison_formatted.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
        except Exception as e:
            st.error(f"An error occurred while processing the files: {e}")